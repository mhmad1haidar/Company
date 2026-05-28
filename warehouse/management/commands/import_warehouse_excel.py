from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from warehouse.models import Category, Item, Supplier


def clean_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def clean_int(value):
    if value in (None, ""):
        return 0
    try:
        return int(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return 0


def clean_decimal(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", ".")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0")


class Command(BaseCommand):
    help = "Import the NIGIT warehouse Excel workbook into warehouse items."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path")
        parser.add_argument("--sheet", default="MagazzinoNigit2026")
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        path = options["xlsx_path"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        try:
            workbook = load_workbook(path, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"File not found: {path}") from exc

        if sheet_name not in workbook.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found. Available: {', '.join(workbook.sheetnames)}")

        sheet = workbook[sheet_name]
        created = 0
        updated = 0
        skipped = 0

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=3, values_only=True):
                code = clean_text(row[0])
                description = clean_text(row[18])
                if not code or code.upper() == "#REF!" or not description:
                    skipped += 1
                    continue

                category_name = clean_text(row[34]) or clean_text(row[35]) or "Uncategorized"
                category, _ = Category.objects.get_or_create(name=category_name)

                supplier_name = clean_text(row[5])
                supplier = None
                if supplier_name:
                    supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

                position_parts = [clean_text(row[index]) for index in (1, 2, 3, 4) if clean_text(row[index])]
                position = clean_text(row[53]) or " ".join(position_parts)

                second_description = clean_text(row[19])
                notes = clean_text(row[32])
                full_description = description
                if second_description:
                    full_description = f"{description}\n{second_description}"

                defaults = {
                    "name": description[:200],
                    "description": full_description,
                    "category": category,
                    "supplier": supplier,
                    "supplier_code": clean_text(row[6]),
                    "unit_of_measure": clean_text(row[20]),
                    "incoming_quantity": clean_int(row[22]),
                    "used_quantity": clean_int(row[23]),
                    "quantity": clean_int(row[24]),
                    "in_transit_quantity": clean_int(row[26]),
                    "unit_price": clean_decimal(row[7]),
                    "status": "out_of_stock" if clean_int(row[24]) == 0 else "active",
                    "location": position[:100],
                    "position_area": clean_text(row[1]),
                    "position_shelf": clean_text(row[2]),
                    "position_level": clean_text(row[3]),
                    "position_detail": clean_text(row[4]),
                    "notes": notes,
                }

                _, was_created = Item.objects.update_or_create(code=code, defaults=defaults)
                if was_created:
                    created += 1
                else:
                    updated += 1

            if dry_run:
                transaction.set_rollback(True)

        action = "Would import" if dry_run else "Imported"
        self.stdout.write(
            self.style.SUCCESS(
                f"{action}: {created} created, {updated} updated, {skipped} skipped from {sheet_name}."
            )
        )

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
from django.contrib import messages
from django.urls import reverse_lazy
from django.db.models import Q, Count, Sum, F, DecimalField, ExpressionWrapper
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.forms import formset_factory
import csv

User = get_user_model()

from .models import Item, Category, Supplier, StockMovement, WarehouseRequest, WarehouseRequestItem, ItemSupplier, WarehouseZone, AuditLog, PurchaseOrder, PurchaseOrderItem, SupplierPerformance, Batch, Stocktaking, StocktakingItem, ItemAssignment, SerialNumber
from .forms import StockMovementForm, WarehouseRequestForm, ItemSupplierForm, WarehouseRequestItemForm, ItemAssignmentForm, SerialNumberForm
from io import BytesIO
import base64
from django.http import HttpResponse


class WarehouseDashboardView(LoginRequiredMixin, ListView):
    """Main warehouse dashboard"""
    model = Item
    template_name = 'warehouse/dashboard.html'
    context_object_name = 'items'

    def get_queryset(self):
        items = Item.objects.filter(status='active')[:10]  # Show recent items
        # Update quantities from stock movements
        for item in items:
            item.quantity = item.calculate_quantity()
        return items

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        active_items = list(Item.objects.filter(status='active'))
        
        # Dashboard statistics
        context.update({
            'total_items': Item.objects.count(),
            'active_items': len(active_items),
            'low_stock_items': sum(1 for item in active_items if item.quantity <= item.min_quantity),
            'out_of_stock_items': sum(1 for item in active_items if item.quantity == 0),
            'in_transit_items': Item.objects.filter(in_transit_quantity__gt=0).count(),
            'stock_movements_count': StockMovement.objects.count(),
            'total_stock_quantity': sum(item.quantity for item in active_items),
            'inventory_value': sum(item.quantity * item.selling_price for item in active_items),
            'negative_stock_items': sum(1 for item in active_items if item.quantity < 0),
            'missing_location_items': Item.objects.filter(Q(location='') | Q(location__isnull=True)).count(),
            'pending_requests': WarehouseRequest.objects.filter(status='pending').count(),
            'total_categories': Category.objects.count(),
            'total_suppliers': Supplier.objects.count(),
            
            # Recent stock movements
            'recent_movements': StockMovement.objects.select_related('item', 'created_by')[:5],
            
            # Low stock items
            'low_stock_items_list': sorted(active_items, key=lambda x: x.quantity)[:5],
            
            # Pending requests
            'pending_requests_list': WarehouseRequest.objects.select_related('requester').prefetch_related('items__item').filter(status='pending')[:5],
            'negative_stock_items_list': sorted([item for item in active_items if item.quantity < 0], key=lambda x: x.quantity)[:5],
            'missing_location_items_list': Item.objects.filter(Q(location='') | Q(location__isnull=True))[:5],
            
            # Recent audit logs (activity feed)
            'recent_audit_logs': AuditLog.objects.select_related('user', 'item').order_by('-created_at')[:10],
            
            # Low stock alerts
            'low_stock_alerts': [item for item in active_items if item.quantity <= item.min_quantity and item.low_stock_alert],
        })
        return context


class ItemListView(LoginRequiredMixin, ListView):
    """List all warehouse items"""
    model = Item
    template_name = 'warehouse/item_list.html'
    context_object_name = 'items'
    paginate_by = 20

    def get_queryset(self):
        queryset = Item.objects.select_related('category').prefetch_related('supplier_details__supplier').all()
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(name__icontains=search) |
                Q(description__icontains=search) |
                Q(barcode__icontains=search) |
                Q(location__icontains=search) |
                Q(notes__icontains=search)
            )
        
        # Filter by category
        category_id = self.request.GET.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)

        # Filter by zone
        zone_id = self.request.GET.get('zone')
        if zone_id:
            queryset = queryset.filter(zone_id=zone_id)

        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            queryset = queryset.filter(supplier_details__supplier_id=supplier_id).distinct()

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by stock status
        stock_status = self.request.GET.get('stock_status')
        if stock_status == 'low_stock':
            queryset = queryset.filter(quantity__lte=F('min_quantity'))
        elif stock_status == 'out_of_stock':
            queryset = queryset.filter(quantity=0)
        elif stock_status == 'in_stock':
            queryset = queryset.filter(quantity__gt=F('min_quantity'))
        elif stock_status == 'negative_stock':
            queryset = queryset.filter(quantity__lt=0)
        elif stock_status == 'in_transit':
            queryset = queryset.filter(in_transit_quantity__gt=0)
        elif stock_status == 'missing_location':
            queryset = queryset.filter(Q(location='') | Q(location__isnull=True))
        
        # Calculate quantities from stock movements
        items_list = list(queryset)
        for item in items_list:
            item.quantity = item.calculate_quantity()
        
        # Sorting
        sort_by = self.request.GET.get('sort', 'code')
        if sort_by == 'code':
            items_list.sort(key=lambda x: x.code)
        elif sort_by == '-code':
            items_list.sort(key=lambda x: x.code, reverse=True)
        elif sort_by == 'name':
            items_list.sort(key=lambda x: x.name)
        elif sort_by == '-name':
            items_list.sort(key=lambda x: x.name, reverse=True)
        elif sort_by == 'quantity':
            items_list.sort(key=lambda x: x.quantity)
        elif sort_by == '-quantity':
            items_list.sort(key=lambda x: x.quantity, reverse=True)
        
        return items_list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'categories': Category.objects.all(),
            'suppliers': Supplier.objects.all(),
            'zones': WarehouseZone.objects.all(),
            'status_choices': Item.STATUS_CHOICES,
        })
        return context


@login_required
def update_item_stock_inline(request, item_id):
    """Update incoming and used quantities directly from the item list."""
    item = get_object_or_404(Item, pk=item_id)
    next_url = request.POST.get('next') or reverse_lazy('warehouse:item_list')

    if request.method == 'POST':
        try:
            incoming_quantity = int(request.POST.get('incoming_quantity', item.incoming_quantity))
            used_quantity = int(request.POST.get('used_quantity', item.used_quantity))
        except (TypeError, ValueError):
            messages.error(request, 'Please enter valid numbers for In and Used.')
        else:
            if incoming_quantity < 0 or used_quantity < 0:
                messages.error(request, 'In and Used cannot be negative.')
            else:
                item.incoming_quantity = incoming_quantity
                item.used_quantity = used_quantity
                item.quantity = item.calculate_quantity()
                item.updated_by = request.user
                item.save(update_fields=['incoming_quantity', 'used_quantity', 'quantity', 'updated_by', 'updated_at'])
                messages.success(request, f'{item.code} stock updated.')

    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse_lazy('warehouse:item_list')
    return redirect(next_url)


class ItemDetailView(LoginRequiredMixin, DetailView):
    """Detail view for a specific item"""
    model = Item
    template_name = 'warehouse/item_detail.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.get_object()
        
        # Calculate quantity from incoming - used and save to database
        item.quantity = item.calculate_quantity()
        item.save(update_fields=['quantity'])
        
        context.update({
            'stock_movements': StockMovement.objects.filter(item=item).select_related('created_by')[:20],
            'related_requests': WarehouseRequest.objects.filter(items__item=item).select_related('requester').distinct()[:10],
            'previous_item': Item.objects.filter(code__lt=item.code).order_by('-code').first(),
            'next_item': Item.objects.filter(code__gt=item.code).order_by('code').first(),
            'supplier_details': item.supplier_details.select_related('supplier').all(),
            'suppliers': Supplier.objects.all(),
            'audit_logs': item.audit_logs.all()[:10],
        })
        return context


class ItemCreateView(LoginRequiredMixin, CreateView):
    """Create a new warehouse item"""
    model = Item
    template_name = 'warehouse/item_form.html'
    fields = ['code', 'name', 'description', 'category', 'barcode', 'zone', 'image',
              'incoming_quantity', 'used_quantity', 'in_transit_quantity',
              'min_quantity', 'max_quantity', 'unit_of_measure', 'selling_price', 'low_stock_alert',
              'status', 'location', 'position_area', 'position_shelf', 'position_level',
              'position_detail', 'notes']

    def get_success_url(self):
        return reverse_lazy('warehouse:item_detail', kwargs={'pk': self.object.pk})

    def get_initial(self):
        initial = super().get_initial()
        # Auto-generate next code based on last mg code
        last_item = Item.objects.filter(code__startswith='mg').order_by('-code').first()
        if last_item and last_item.code:
            try:
                # Extract number from code like "mg0015" -> 15
                code_num = int(last_item.code[2:])
                next_num = code_num + 1
                initial['code'] = f"mg{next_num:04d}"  # Pad to 4 digits
            except (ValueError, IndexError):
                initial['code'] = "mg0001"
        else:
            initial['code'] = "mg0001"
        
        # Auto-generate barcode if not provided
        if not initial.get('barcode'):
            initial['barcode'] = f"BC{initial['code']}"
        
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        SupplierFormSet = formset_factory(ItemSupplierForm, extra=1, can_delete=True, validate_min=False)
        AssignmentFormSet = formset_factory(ItemAssignmentForm, extra=1, can_delete=True, validate_min=False)
        SerialFormSet = formset_factory(SerialNumberForm, extra=1, can_delete=True, validate_min=False)
        if 'supplier_formset' not in context:
            context['supplier_formset'] = SupplierFormSet()
        if 'assignment_formset' not in context:
            context['assignment_formset'] = AssignmentFormSet()
        if 'serial_formset' not in context:
            context['serial_formset'] = SerialFormSet()
        context.update({
            'categories': Category.objects.all(),
            'suppliers': Supplier.objects.all(),
            'zones': WarehouseZone.objects.all(),
            'title': 'Add New Item',
            'users': User.objects.all().order_by('username'),
        })
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        SupplierFormSet = formset_factory(ItemSupplierForm, extra=1, can_delete=True, validate_min=False)
        AssignmentFormSet = formset_factory(ItemAssignmentForm, extra=1, can_delete=True, validate_min=False)
        SerialFormSet = formset_factory(SerialNumberForm, extra=1, can_delete=True, validate_min=False)
        supplier_formset = SupplierFormSet(request.POST)
        assignment_formset = AssignmentFormSet(request.POST)
        serial_formset = SerialFormSet(request.POST)
        
        # Validate main form and formsets
        formsets_valid = True
        if not supplier_formset.is_valid():
            formsets_valid = False
        if not assignment_formset.is_valid():
            formsets_valid = False
        if not serial_formset.is_valid():
            formsets_valid = False
        
        if form.is_valid() and formsets_valid:
            return self.form_valid(form, supplier_formset, assignment_formset, serial_formset)
        else:
            return self.form_invalid(form, supplier_formset, assignment_formset, serial_formset)

    def form_valid(self, form, supplier_formset, assignment_formset, serial_formset):
        form.instance.created_by = self.request.user
        self.object = form.save()
        
        # Create audit log
        AuditLog.objects.create(
            item=self.object,
            user=self.request.user,
            action='create',
            new_value=f"Item {self.object.code} - {self.object.name} created"
        )
        
        # Save suppliers
        for supplier_form in supplier_formset:
            if supplier_form.is_valid() and not supplier_form.cleaned_data.get('DELETE'):
                # Only process if a supplier is actually selected
                if supplier_form.cleaned_data.get('supplier'):
                    supplier = supplier_form.save(commit=False)
                    supplier.item = self.object
                    supplier.save()
        
        # Save assignments
        for assignment_form in assignment_formset:
            if assignment_form.is_valid() and not assignment_form.cleaned_data.get('DELETE'):
                # Only process if a user is actually selected
                if assignment_form.cleaned_data.get('user'):
                    assignment = assignment_form.save(commit=False)
                    assignment.item = self.object
                    assignment.assigned_by = self.request.user
                    assignment.save()
        
        # Save serial numbers
        for serial_form in serial_formset:
            if not serial_form.cleaned_data.get('DELETE'):
                # Only process if a serial number is actually provided
                if serial_form.cleaned_data.get('serial_number'):
                    try:
                        serial = serial_form.save(commit=False)
                        serial.item = self.object
                        if serial.assigned_to:
                            serial.assigned_at = timezone.now()
                        serial.save()
                    except Exception as e:
                        messages.error(self.request, f'Error saving serial number: {e}')
        
        messages.success(self.request, f'Item "{form.instance.name}" has been created successfully with suppliers, assignments, and serial numbers.')
        return redirect(self.get_success_url())

    def form_invalid(self, form, supplier_formset, assignment_formset, serial_formset):
        return self.render_to_response(self.get_context_data(form=form, supplier_formset=supplier_formset, assignment_formset=assignment_formset, serial_formset=serial_formset))


class ItemUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing warehouse item"""
    model = Item
    template_name = 'warehouse/item_form.html'
    fields = ['code', 'name', 'description', 'category', 'barcode', 'zone', 'image',
              'incoming_quantity', 'used_quantity', 'in_transit_quantity',
              'min_quantity', 'max_quantity', 'unit_of_measure', 'selling_price', 'low_stock_alert',
              'status', 'location', 'position_area', 'position_shelf', 'position_level',
              'position_detail', 'notes']
    success_url = reverse_lazy('warehouse:item_list')

    def form_valid(self, form):
        form.instance.updated_by = self.request.user
        messages.success(self.request, f'Item "{form.instance.name}" has been updated successfully.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f'{field}: {error}')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        SupplierFormSet = formset_factory(ItemSupplierForm, extra=1, can_delete=True, validate_min=False)
        AssignmentFormSet = formset_factory(ItemAssignmentForm, extra=1, can_delete=True, validate_min=False)
        SerialFormSet = formset_factory(SerialNumberForm, extra=1, can_delete=True, validate_min=False)
        if 'supplier_formset' not in context:
            context['supplier_formset'] = SupplierFormSet()
        if 'assignment_formset' not in context:
            context['assignment_formset'] = AssignmentFormSet()
        if 'serial_formset' not in context:
            context['serial_formset'] = SerialFormSet()
        context.update({
            'categories': Category.objects.all(),
            'suppliers': Supplier.objects.all(),
            'zones': WarehouseZone.objects.all(),
            'title': 'Edit Item',
            'supplier_details': self.object.supplier_details.select_related('supplier').all() if self.object.pk else [],
            'users': User.objects.all().order_by('username'),
        })
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form()
        SupplierFormSet = formset_factory(ItemSupplierForm, extra=1, can_delete=True, validate_min=False)
        AssignmentFormSet = formset_factory(ItemAssignmentForm, extra=1, can_delete=True, validate_min=False)
        SerialFormSet = formset_factory(SerialNumberForm, extra=1, can_delete=True, validate_min=False)
        supplier_formset = SupplierFormSet(request.POST)
        assignment_formset = AssignmentFormSet(request.POST)
        serial_formset = SerialFormSet(request.POST)
        
        # Validate main form and formsets
        formsets_valid = True
        if not supplier_formset.is_valid():
            formsets_valid = False
        if not assignment_formset.is_valid():
            formsets_valid = False
        if not serial_formset.is_valid():
            formsets_valid = False
        
        if form.is_valid() and formsets_valid:
            return self.form_valid(form, supplier_formset, assignment_formset, serial_formset)
        else:
            return self.form_invalid(form, supplier_formset, assignment_formset, serial_formset)

    def form_valid(self, form, supplier_formset, assignment_formset, serial_formset):
        form.instance.updated_by = self.request.user
        self.object = form.save()
        
        # Save suppliers
        for supplier_form in supplier_formset:
            if not supplier_form.cleaned_data.get('DELETE') and supplier_form.is_valid():
                # Only process if a supplier is actually selected
                if supplier_form.cleaned_data.get('supplier'):
                    supplier = supplier_form.save(commit=False)
                    supplier.item = self.object
                    supplier.save()
        
        # Save assignments
        for assignment_form in assignment_formset:
            if not assignment_form.cleaned_data.get('DELETE') and assignment_form.is_valid():
                # Only process if a user is actually selected
                if assignment_form.cleaned_data.get('user'):
                    assignment = assignment_form.save(commit=False)
                    assignment.item = self.object
                    assignment.assigned_by = self.request.user
                    assignment.save()
        
        # Save serial numbers
        for serial_form in serial_formset:
            if not serial_form.cleaned_data.get('DELETE'):
                # Only process if a serial number is actually provided
                if serial_form.cleaned_data.get('serial_number'):
                    try:
                        serial = serial_form.save(commit=False)
                        serial.item = self.object
                        if serial.assigned_to:
                            serial.assigned_at = timezone.now()
                        serial.save()
                    except Exception as e:
                        messages.error(self.request, f'Error saving serial number: {e}')
        
        messages.success(self.request, f'Item "{form.instance.name}" has been updated successfully.')
        return redirect(self.get_success_url())
    
    def form_invalid(self, form, supplier_formset, assignment_formset, serial_formset):
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f'{field}: {error}')
        return self.render_to_response(self.get_context_data(form=form, supplier_formset=supplier_formset, assignment_formset=assignment_formset, serial_formset=serial_formset))


class AddItemSupplierView(LoginRequiredMixin, CreateView):
    """Add a supplier relationship to an item"""
    model = ItemSupplier
    form_class = ItemSupplierForm
    template_name = 'warehouse/item_detail.html'

    def dispatch(self, request, *args, **kwargs):
        self.item = get_object_or_404(Item, pk=kwargs['item_id'])
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.item = self.item
        messages.success(self.request, f'Supplier "{form.instance.supplier.name}" has been added to item "{self.item.code}" with price {form.instance.unit_price}.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('warehouse:item_detail', kwargs={'pk': self.item.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.item
        context['supplier_details'] = self.item.supplier_details.select_related('supplier').all()
        context['stock_movements'] = StockMovement.objects.filter(item=self.item).select_related('created_by')[:20]
        context['related_requests'] = WarehouseRequest.objects.filter(item=self.item).select_related('requester')[:10]
        context['previous_item'] = Item.objects.filter(code__lt=self.item.code).order_by('-code').first()
        context['next_item'] = Item.objects.filter(code__gt=self.item.code).order_by('code').first()
        context['add_supplier_form'] = self.get_form()
        return context


class EditItemSupplierView(LoginRequiredMixin, UpdateView):
    """Edit a supplier relationship"""
    model = ItemSupplier
    form_class = ItemSupplierForm
    template_name = 'warehouse/edit_supplier.html'
    context_object_name = 'supplier_detail'

    def get_success_url(self):
        return reverse_lazy('warehouse:item_detail', kwargs={'pk': self.object.item.pk})

    def form_valid(self, form):
        messages.success(self.request, f'Supplier "{form.instance.supplier.name}" has been updated.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['item'] = self.object.item
        context['suppliers'] = Supplier.objects.all()
        return context


@login_required
def delete_item_supplier(request, pk):
    """Delete a supplier relationship"""
    supplier_detail = get_object_or_404(ItemSupplier, pk=pk)
    item_pk = supplier_detail.item.pk
    supplier_name = supplier_detail.supplier.name
    
    if request.method == 'POST':
        supplier_detail.delete()
        messages.success(request, f'Supplier "{supplier_name}" has been removed.')
        return redirect('warehouse:item_detail', pk=item_pk)
    
    return render(request, 'warehouse/delete_supplier.html', {
        'supplier_detail': supplier_detail,
        'item': supplier_detail.item
    })


@login_required
def create_supplier_ajax(request):
    """Create a new supplier via AJAX"""
    if request.method == 'POST':
        name = request.POST.get('name')
        contact_person = request.POST.get('contact_person', '')
        email = request.POST.get('email', '')
        phone = request.POST.get('phone', '')
        address = request.POST.get('address', '')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Supplier name is required'})
        
        try:
            supplier = Supplier.objects.create(
                name=name,
                contact_person=contact_person,
                email=email,
                phone=phone,
                address=address
            )
            return JsonResponse({
                'success': True,
                'supplier_id': supplier.pk,
                'supplier_name': supplier.name
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


class StockMovementListView(LoginRequiredMixin, ListView):
    """List all stock movements"""
    model = StockMovement
    template_name = 'warehouse/stock_movement_list.html'
    context_object_name = 'movements'
    paginate_by = 30

    def get_queryset(self):
        queryset = StockMovement.objects.select_related('item', 'created_by').all()
        
        # Filter by item
        item_id = self.request.GET.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)
        
        # Filter by movement type
        movement_type = self.request.GET.get('movement_type')
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(created_at__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__date__lte=end_date)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'items': Item.objects.all(),
            'movement_types': StockMovement.MOVEMENT_TYPES,
        })
        return context


class StockMovementCreateView(LoginRequiredMixin, FormView):
    template_name = 'warehouse/stock_movement_form.html'
    form_class = StockMovementForm
    success_url = reverse_lazy('warehouse:movement_list')

    def get_initial(self):
        initial = super().get_initial()
        item_id = self.request.GET.get('item')
        if item_id:
            initial['item'] = item_id
        return initial

    def form_valid(self, form):
        item = form.cleaned_data['item']
        movement_type = form.cleaned_data['movement_type']
        entered_quantity = form.cleaned_data['quantity']
        previous_quantity = item.quantity

        if movement_type == 'in':
            movement_quantity = entered_quantity
            new_quantity = previous_quantity + entered_quantity
        elif movement_type == 'out':
            if entered_quantity > previous_quantity:
                form.add_error('quantity', 'Not enough stock available for this outgoing movement.')
                return self.form_invalid(form)
            movement_quantity = -entered_quantity
            new_quantity = previous_quantity - entered_quantity
        elif movement_type == 'adjustment':
            new_quantity = entered_quantity
            movement_quantity = new_quantity - previous_quantity
        else:
            movement_quantity = 0
            new_quantity = previous_quantity

        StockMovement.objects.create(
            item=item,
            movement_type=movement_type,
            quantity=movement_quantity,
            previous_quantity=previous_quantity,
            new_quantity=new_quantity,
            reason=form.cleaned_data['reason'],
            notes=form.cleaned_data['notes'],
            created_by=self.request.user
        )
        
        # Update incoming/used quantities based on movement type
        if movement_type == 'in':
            item.incoming_quantity += entered_quantity
        elif movement_type == 'out':
            item.used_quantity += entered_quantity
        
        # Calculate quantity from incoming - used
        item.quantity = item.calculate_quantity()
        item.save()
        
        messages.success(self.request, f'Stock movement recorded successfully. New quantity: {item.quantity}')
        return super().form_valid(form)


class WarehouseRequestListView(LoginRequiredMixin, ListView):
    """List warehouse requests"""
    model = WarehouseRequest
    template_name = 'warehouse/request_list.html'
    context_object_name = 'requests'
    paginate_by = 20

    def get_queryset(self):
        queryset = WarehouseRequest.objects.select_related('requester').prefetch_related('items__item').all()
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by requester (for non-admin users)
        if not self.request.user.is_staff:
            queryset = queryset.filter(requester=self.request.user)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'status_choices': WarehouseRequest.STATUS_CHOICES,
        })
        return context


class WarehouseRequestCreateView(LoginRequiredMixin, CreateView):
    model = WarehouseRequest
    form_class = WarehouseRequestForm
    template_name = 'warehouse/request_form.html'
    success_url = reverse_lazy('warehouse:request_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        WarehouseRequestItemFormSet = formset_factory(WarehouseRequestItemForm, extra=1, can_delete=True)
        if 'item_formset' not in context:
            context['item_formset'] = WarehouseRequestItemFormSet()
        context.update({
            'items': Item.objects.filter(status='active').order_by('code', 'name'),
        })
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        WarehouseRequestItemFormSet = formset_factory(WarehouseRequestItemForm, extra=1, can_delete=True)
        item_formset = WarehouseRequestItemFormSet(request.POST)
        
        if form.is_valid() and item_formset.is_valid():
            return self.form_valid(form, item_formset)
        else:
            return self.form_invalid(form, item_formset)

    def form_valid(self, form, item_formset):
        form.instance.requester = self.request.user
        self.object = form.save()
        
        # Save items
        for item_form in item_formset:
            if not item_form.cleaned_data.get('DELETE'):
                item = item_form.save(commit=False)
                item.request = self.object
                item.save()
        
        messages.success(self.request, f'Warehouse request created successfully with {len([f for f in item_formset if not f.cleaned_data.get('DELETE')])} item(s).')
        return super().form_valid(form)


@login_required
def approve_request(request, pk):
    """Approve a warehouse request"""
    warehouse_request = get_object_or_404(WarehouseRequest, pk=pk)
    
    if request.method == 'POST':
        # Approve all items in the request
        for request_item in warehouse_request.items.all():
            request_item.quantity_approved = request_item.quantity_requested
            request_item.save()
        
        warehouse_request.status = 'approved'
        warehouse_request.approved_by = request.user
        warehouse_request.approved_at = timezone.now()
        warehouse_request.save()
        
        messages.success(request, f'Request #{warehouse_request.id} approved successfully.')
    
    return redirect('warehouse:request_list')


@login_required
def reject_request(request, pk):
    """Reject a warehouse request"""
    warehouse_request = get_object_or_404(WarehouseRequest, pk=pk)
    
    if request.method == 'POST':
        warehouse_request.status = 'rejected'
        warehouse_request.approved_by = request.user
        warehouse_request.approved_at = timezone.now()
        warehouse_request.save()
        
        messages.success(request, f'Request #{warehouse_request.id} rejected successfully.')
    
    return redirect('warehouse:request_list')


@login_required
def generate_barcode_image(request, item_id):
    """Generate barcode image for an item"""
    try:
        import barcode
        from barcode.writer import ImageWriter
    except ImportError:
        return HttpResponse("python-barcode library not installed", status=500)

    item = get_object_or_404(Item, pk=item_id)
    barcode_code = item.barcode or item.code
    
    # Generate barcode
    barcode_class = barcode.get_barcode_class('code128')
    barcode_instance = barcode_class(barcode_code, writer=ImageWriter())
    
    # Generate PNG image
    buffer = BytesIO()
    barcode_instance.write(buffer, options={'module_height': 10, 'font_size': 10})
    buffer.seek(0)
    
    return HttpResponse(buffer.getvalue(), content_type='image/png')


@login_required
def generate_qr_code(request, item_id):
    """Generate QR code for an item"""
    try:
        import qrcode
    except ImportError:
        return HttpResponse("qrcode library not installed", status=500)

    item = get_object_or_404(Item, pk=item_id)
    qr_data = f"{item.code}|{item.name}|{item.barcode}"
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Generate PNG image
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    return HttpResponse(buffer.getvalue(), content_type='image/png')


@login_required
def inventory_report(request):
    """Generate inventory valuation report"""
    items = Item.objects.select_related('category', 'zone').all()
    total_value = sum(item.quantity * item.selling_price for item in items)
    
    # Add calculated value to each item
    items_with_value = []
    for item in items:
        items_with_value.append({
            'item': item,
            'value': item.quantity * item.selling_price
        })
    
    return render(request, 'warehouse/inventory_report.html', {
        'items_with_value': items_with_value,
        'total_value': total_value,
    })


@login_required
def print_barcode_labels(request):
    """Print barcode labels for selected items"""
    item_ids = request.GET.getlist('items')
    items = Item.objects.filter(pk__in=item_ids)
    
    return render(request, 'warehouse/print_labels.html', {
        'items': items,
    })


class PurchaseOrderListView(LoginRequiredMixin, ListView):
    """List all purchase orders"""
    model = PurchaseOrder
    template_name = 'warehouse/purchase_order_list.html'
    context_object_name = 'orders'
    paginate_by = 20


class PurchaseOrderDetailView(LoginRequiredMixin, DetailView):
    """Detail view for a purchase order"""
    model = PurchaseOrder
    template_name = 'warehouse/purchase_order_detail.html'
    context_object_name = 'order'


class PurchaseOrderCreateView(LoginRequiredMixin, CreateView):
    """Create a new purchase order"""
    model = PurchaseOrder
    template_name = 'warehouse/purchase_order_form.html'
    fields = ['order_number', 'supplier', 'expected_delivery_date', 'notes']

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        return super().form_valid(form)


class PurchaseOrderUpdateView(LoginRequiredMixin, UpdateView):
    """Update a purchase order"""
    model = PurchaseOrder
    template_name = 'warehouse/purchase_order_form.html'
    fields = ['order_number', 'supplier', 'expected_delivery_date', 'status', 'notes']
    success_url = reverse_lazy('warehouse:purchase_order_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Auto-calculate supplier performance
        supplier = form.instance.supplier
        orders = PurchaseOrder.objects.filter(supplier=supplier)
        
        total_orders = orders.count()
        on_time_deliveries = orders.filter(expected_delivery_date__gte=F('order_date')).count()
        late_deliveries = total_orders - on_time_deliveries
        
        performance, created = SupplierPerformance.objects.get_or_create(
            supplier=supplier
        )
        performance.total_orders = total_orders
        performance.on_time_deliveries = on_time_deliveries
        performance.late_deliveries = late_deliveries
        performance.last_order_date = form.instance.order_date
        performance.total_spent = orders.aggregate(total=Sum('items__quantity_ordered' * F('items__unit_price')))['total'] or 0
        performance.save()
        
        return response


class BatchListView(LoginRequiredMixin, ListView):
    """List all batches"""
    model = Batch
    template_name = 'warehouse/batch_list.html'
    context_object_name = 'batches'
    paginate_by = 20


class BatchDetailView(LoginRequiredMixin, DetailView):
    """Detail view for a batch"""
    model = Batch
    template_name = 'warehouse/batch_detail.html'
    context_object_name = 'batch'


class BatchCreateView(LoginRequiredMixin, CreateView):
    """Create a new batch"""
    model = Batch
    template_name = 'warehouse/batch_form.html'
    fields = ['item', 'batch_number', 'quantity', 'manufacturing_date', 'expiration_date', 'supplier', 'notes']
    success_url = reverse_lazy('warehouse:batch_list')


class BatchUpdateView(LoginRequiredMixin, UpdateView):
    """Update a batch"""
    model = Batch
    template_name = 'warehouse/batch_form.html'
    fields = ['item', 'batch_number', 'quantity', 'manufacturing_date', 'expiration_date', 'supplier', 'notes']
    success_url = reverse_lazy('warehouse:batch_list')


class StocktakingListView(LoginRequiredMixin, ListView):
    """List all stocktakings"""
    model = Stocktaking
    template_name = 'warehouse/stocktaking_list.html'
    context_object_name = 'stocktakings'
    paginate_by = 20


class StocktakingDetailView(LoginRequiredMixin, DetailView):
    """Detail view for a stocktaking"""
    model = Stocktaking
    template_name = 'warehouse/stocktaking_detail.html'
    context_object_name = 'stocktaking'


class StocktakingCreateView(LoginRequiredMixin, CreateView):
    """Create a new stocktaking"""
    model = Stocktaking
    template_name = 'warehouse/stocktaking_form.html'
    fields = ['name', 'description', 'notes']
    
    def form_valid(self, form):
        form.instance.conducted_by = self.request.user
        return super().form_valid(form)


class UserAssignmentsListView(LoginRequiredMixin, ListView):
    """List all user assignments"""
    model = ItemAssignment
    template_name = 'warehouse/assignments_list.html'
    context_object_name = 'assignments'
    paginate_by = 20

    def get_queryset(self):
        queryset = ItemAssignment.objects.select_related('item', 'user', 'assigned_by').all()
        
        # Filter by user
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by item
        item_id = self.request.GET.get('item')
        if item_id:
            queryset = queryset.filter(item_id=item_id)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get serial numbers with filtering
        serial_numbers = SerialNumber.objects.select_related('item', 'assigned_to').all()
        
        # Filter by search
        search = self.request.GET.get('search')
        if search:
            serial_numbers = serial_numbers.filter(serial_number__icontains=search)
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            serial_numbers = serial_numbers.filter(status=status)
        
        # Filter by assigned user
        user_id = self.request.GET.get('user')
        if user_id:
            serial_numbers = serial_numbers.filter(assigned_to_id=user_id)
        
        serial_numbers = serial_numbers.order_by('-created_at')
        
        context.update({
            'users': User.objects.all().order_by('username'),
            'items': Item.objects.filter(status='active').order_by('code', 'name'),
            'serial_numbers': serial_numbers,
        })
        return context


@login_required
def assign_serial(request, pk):
    """Assign a serial number to a user"""
    serial = get_object_or_404(SerialNumber, pk=pk)
    
    if request.method == 'POST':
        assigned_to_id = request.POST.get('assigned_to')
        if assigned_to_id:
            assigned_to = get_object_or_404(User, pk=assigned_to_id)
            serial.assigned_to = assigned_to
            serial.status = 'assigned'
            serial.assigned_at = timezone.now()
            serial.save()
            full_name = assigned_to.get_full_name or assigned_to.username
            messages.success(request, f'Serial number {serial.serial_number} assigned to {full_name}')
        else:
            messages.error(request, 'Please select a user to assign the serial number to')
    
    return redirect('warehouse:assignments_list')


@login_required
def unassign_serial(request, pk):
    """Unassign a serial number from a user"""
    serial = get_object_or_404(SerialNumber, pk=pk)
    
    if request.method == 'POST':
        serial.assigned_to = None
        serial.status = 'available'
        serial.assigned_at = None
        serial.save()
        messages.success(request, f'Serial number {serial.serial_number} unassigned successfully')
    
    return redirect('warehouse:assignments_list')


@login_required
def export_items_csv(request):
    """Export items to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="items.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Code', 'Name', 'Description', 'Category', 'Quantity', 'Unit', 'Price', 'Status', 'Location'])
    
    items = Item.objects.select_related('category').all()
    for item in items:
        writer.writerow([
            item.code,
            item.name,
            item.description,
            item.category.name if item.category else '',
            item.calculate_quantity(),
            item.unit_of_measure,
            item.selling_price,
            item.status,
            item.location
        ])
    
    return response


def _filtered_assignments(request):
    assignments = ItemAssignment.objects.select_related('item', 'user', 'assigned_by').all()
    user_id = request.GET.get('user')
    item_id = request.GET.get('item')
    if user_id:
        assignments = assignments.filter(user_id=user_id)
    if item_id:
        assignments = assignments.filter(item_id=item_id)
    return assignments.order_by('-created_at')


def _filtered_serial_numbers(request):
    serials = SerialNumber.objects.select_related('item', 'assigned_to').all()
    search = request.GET.get('search')
    status = request.GET.get('status')
    user_id = request.GET.get('user')
    if search:
        serials = serials.filter(serial_number__icontains=search)
    if status:
        serials = serials.filter(status=status)
    if user_id:
        serials = serials.filter(assigned_to_id=user_id)
    return serials.order_by('-created_at')


def _assignment_export_rows(assignments):
    rows = [['User', 'Item Code', 'Item Name', 'Quantity', 'Notes', 'Assigned By', 'Date']]
    for assignment in assignments:
        rows.append([
            assignment.user.get_full_name() or assignment.user.username,
            assignment.item.code,
            assignment.item.name,
            assignment.quantity_assigned,
            assignment.notes or '',
            assignment.assigned_by.get_full_name() or assignment.assigned_by.username if assignment.assigned_by else '',
            assignment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    return rows


def _serial_export_rows(serials):
    rows = [['Serial Number', 'Item Code', 'Item Name', 'Status', 'Assigned To', 'Assigned At', 'Notes']]
    for serial in serials:
        rows.append([
            serial.serial_number,
            serial.item.code,
            serial.item.name,
            serial.get_status_display(),
            serial.assigned_to.get_full_name() or serial.assigned_to.username if serial.assigned_to else '',
            serial.assigned_at.strftime('%Y-%m-%d %H:%M') if serial.assigned_at else '',
            serial.notes or '',
        ])
    return rows


def _combined_assignment_export_rows(request):
    rows = [[
        'Record Type', 'User / Assigned To', 'Item Code', 'Item Name', 'Quantity',
        'Serial Number', 'Status', 'Notes', 'Handled By', 'Date'
    ]]
    for assignment in _filtered_assignments(request):
        rows.append([
            'Item Assignment',
            assignment.user.get_full_name() or assignment.user.username,
            assignment.item.code,
            assignment.item.name,
            assignment.quantity_assigned,
            '',
            '',
            assignment.notes or '',
            assignment.assigned_by.get_full_name() or assignment.assigned_by.username if assignment.assigned_by else '',
            assignment.created_at.strftime('%Y-%m-%d %H:%M'),
        ])
    for serial in _filtered_serial_numbers(request):
        rows.append([
            'Serial Number',
            serial.assigned_to.get_full_name() or serial.assigned_to.username if serial.assigned_to else '',
            serial.item.code,
            serial.item.name,
            '',
            serial.serial_number,
            serial.get_status_display(),
            serial.notes or '',
            '',
            serial.assigned_at.strftime('%Y-%m-%d %H:%M') if serial.assigned_at else '',
        ])
    return rows


def _export_rows_csv(rows, filename):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerows(rows)
    return response


def _export_rows_excel(rows, filename, title):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl library not installed", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    for cell in ws[1]:
        cell.font = Font(color='FFFFFF', bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 3, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def _export_rows_pdf(rows, filename, title):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse("reportlab library not installed", status=500)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.45*inch, bottomMargin=0.45*inch)
    styles = getSampleStyleSheet()
    table_rows = rows[:300]
    story = [Paragraph(title, styles['Title']), Spacer(1, 0.2*inch)]
    cell_style = styles['BodyText']
    cell_style.fontSize = 7
    cell_style.leading = 8
    wrapped_rows = [
        [Paragraph(str(value or ''), cell_style) for value in row]
        for row in table_rows
    ]
    usable_width = landscape(letter)[0] - doc.leftMargin - doc.rightMargin
    column_count = len(wrapped_rows[0]) if wrapped_rows else 1
    if column_count >= 10:
        col_widths = [
            0.82*inch, 1.05*inch, 0.72*inch, 1.55*inch, 0.55*inch,
            0.95*inch, 0.75*inch, 1.15*inch, 0.95*inch, 0.82*inch
        ]
    else:
        col_widths = [usable_width / column_count] * column_count
    table = Table(wrapped_rows, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#d1d5db')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)
    if len(rows) > 300:
        story.append(Spacer(1, 0.18*inch))
        story.append(Paragraph('PDF preview limited to first 300 rows. Use Excel or CSV for the full export.', styles['Normal']))
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def _export_rows(rows, filename, title, file_format):
    if file_format == 'excel':
        return _export_rows_excel(rows, filename, title)
    if file_format == 'pdf':
        return _export_rows_pdf(rows, filename, title)
    return _export_rows_csv(rows, filename)


@login_required
def export_assignments(request, file_format='csv'):
    """Export assignments to CSV, Excel, or PDF."""
    rows = _assignment_export_rows(_filtered_assignments(request))
    return _export_rows(rows, 'warehouse_assignments', 'Warehouse Assignments', file_format)


@login_required
def export_serial_numbers(request, file_format='csv'):
    """Export serial numbers to CSV, Excel, or PDF."""
    rows = _serial_export_rows(_filtered_serial_numbers(request))
    return _export_rows(rows, 'warehouse_serial_numbers', 'Warehouse Serial Numbers', file_format)


@login_required
def export_assignment_overview(request, file_format='csv'):
    """Export assignments and serial numbers together in one file."""
    rows = _combined_assignment_export_rows(request)
    return _export_rows(rows, 'warehouse_assignments_overview', 'Warehouse Assignments Overview', file_format)


@login_required
def export_assignments_csv(request):
    return export_assignments(request, 'csv')


@login_required
def export_serial_numbers_csv(request):
    return export_serial_numbers(request, 'csv')


@login_required
def bulk_delete_items(request):
    """Bulk delete selected items"""
    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids')
        if item_ids:
            Item.objects.filter(id__in=item_ids).delete()
            messages.success(request, f'{len(item_ids)} items deleted successfully')
    return redirect('warehouse:item_list')


@login_required
def bulk_update_items(request):
    """Bulk update selected warehouse item fields."""
    next_url = request.POST.get('next') or reverse_lazy('warehouse:item_list')

    if request.method == 'POST':
        item_ids = request.POST.getlist('item_ids')
        field_name = request.POST.get('bulk_field')
        value = request.POST.get('bulk_value', '').strip()

        allowed_fields = {'location', 'category', 'zone', 'status'}
        if not item_ids:
            messages.error(request, 'Select at least one item to update.')
        elif field_name not in allowed_fields:
            messages.error(request, 'Choose a valid field to update.')
        else:
            items = Item.objects.filter(id__in=item_ids)
            update_data = {'updated_by': request.user}

            if field_name == 'location':
                update_data['location'] = value
            elif field_name == 'category':
                update_data['category'] = Category.objects.filter(pk=value).first() if value else None
            elif field_name == 'zone':
                update_data['zone'] = WarehouseZone.objects.filter(pk=value).first() if value else None
            elif field_name == 'status':
                valid_statuses = {choice[0] for choice in Item.STATUS_CHOICES}
                if value not in valid_statuses:
                    messages.error(request, 'Choose a valid status.')
                    items = Item.objects.none()
                else:
                    update_data['status'] = value

            updated = 0
            for item in items:
                for key, field_value in update_data.items():
                    setattr(item, key, field_value)
                item.save()
                updated += 1

            if updated:
                messages.success(request, f'{updated} item(s) updated successfully.')

    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse_lazy('warehouse:item_list')
    return redirect(next_url)


@login_required
def generate_barcode(request, item_id):
    """Generate barcode for an item"""
    item = get_object_or_404(Item, pk=item_id)
    from io import BytesIO
    from reportlab.lib.pagesizes import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="barcode_{item.code}.pdf"'
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=(2*inch, 1*inch))
    
    # Draw barcode
    p.setFont("Helvetica-Bold", 12)
    p.drawString(10, 50, item.code)
    p.drawString(10, 35, item.name[:30])
    
    p.save()
    buffer.seek(0)
    response.write(buffer.read())
    buffer.close()
    
    return response


@login_required
def export_inventory_pdf(request):
    """Export inventory report to PDF"""
    from io import BytesIO
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Create a custom style for wrapped text in table cells
    wrapped_style = ParagraphStyle(
        'wrapped',
        parent=styles['Normal'],
        fontSize=7,
        alignment=1  # CENTER
    )
    
    # Title
    elements.append(Paragraph("Warehouse Inventory Report", styles['Title']))
    elements.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Get items data
    items = Item.objects.select_related('category').prefetch_related('supplier_details__supplier').all()
    data = [['Code', 'Name', 'Category', 'Suppliers', 'Selling Price', 'Quantity', 'Status', 'Location']]
    for item in items:
        suppliers = ', '.join([s.supplier.name for s in item.supplier_details.all()]) if item.supplier_details.exists() else '-'
        data.append([
            Paragraph(str(item.code), wrapped_style),
            Paragraph(str(item.name)[:40], wrapped_style),
            Paragraph(str(item.category.name if item.category else ''), wrapped_style),
            Paragraph(str(suppliers)[:35], wrapped_style),
            Paragraph(f"${item.selling_price:.2f}" if item.selling_price else "-", wrapped_style),
            Paragraph(str(item.calculate_quantity()), wrapped_style),
            Paragraph(str(item.get_status_display()), wrapped_style),
            Paragraph(str(item.location or '')[:30], wrapped_style)
        ])
    
    # Create table with column widths
    col_widths = [0.9*inch, 1.5*inch, 1.1*inch, 1.3*inch, 0.8*inch, 0.8*inch, 0.9*inch, 1*inch]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response.write(buffer.read())
    buffer.close()
    
    return response


@login_required
def export_inventory_csv(request):
    """Export inventory report to CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Code', 'Name', 'Category', 'Suppliers', 'Selling Price', 'Quantity', 'Status', 'Location'])
    
    items = Item.objects.select_related('category').prefetch_related('supplier_details__supplier').all()
    for item in items:
        suppliers = ', '.join([s.supplier.name for s in item.supplier_details.all()]) if item.supplier_details.exists() else '-'
        writer.writerow([
            item.code,
            item.name,
            item.category.name if item.category else '',
            suppliers,
            f"${item.selling_price:.2f}" if item.selling_price else "-",
            item.calculate_quantity(),
            item.get_status_display(),
            item.location or ''
        ])
    
    return response


@login_required
def export_inventory_excel(request):
    """Export inventory report to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl library not installed", status=500)
    
    from io import BytesIO
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventory_report.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    headers = ['Code', 'Name', 'Category', 'Suppliers', 'Selling Price', 'Quantity', 'Status', 'Location']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    items = Item.objects.select_related('category').prefetch_related('supplier_details__supplier').all()
    for item in items:
        suppliers = ', '.join([s.supplier.name for s in item.supplier_details.all()]) if item.supplier_details.exists() else '-'
        ws.append([
            item.code,
            item.name,
            item.category.name if item.category else '',
            suppliers,
            f"${item.selling_price:.2f}" if item.selling_price else "-",
            item.calculate_quantity(),
            item.get_status_display(),
            item.location or ''
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    
    return response


@login_required
def generate_qr_code_pdf(request, item_id):
    """Generate QR code for an item"""
    item = get_object_or_404(Item, pk=item_id)
    from io import BytesIO
    from reportlab.lib.pagesizes import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    try:
        import qrcode
    except ImportError:
        return HttpResponse("qrcode library not installed", status=500)
    from PIL import Image
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="qr_{item.code}.pdf"'
    
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=(2*inch, 2.5*inch))
    
    # Create QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(f"{item.code}|{item.name}|{item.calculate_quantity()}")
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save QR code to temporary buffer
    qr_buffer = BytesIO()
    img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    
    # Draw QR code on PDF
    from reportlab.lib.utils import ImageReader
    p.drawImage(ImageReader(qr_buffer), 0.25*inch, 0.8*inch, width=1.5*inch, height=1.5*inch)
    
    # Draw item info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(10, 50, item.code)
    p.setFont("Helvetica", 10)
    p.drawString(10, 35, item.name[:30])
    p.drawString(10, 20, f"Qty: {item.calculate_quantity()}")
    
    p.save()
    buffer.seek(0)
    response.write(buffer.read())
    buffer.close()
    
    return response


class AuditLogListView(LoginRequiredMixin, ListView):
    """View all audit logs"""
    model = AuditLog
    template_name = 'warehouse/audit_log_list.html'
    context_object_name = 'audit_logs'
    paginate_by = 50

    def get_queryset(self):
        queryset = AuditLog.objects.select_related('user').all()
        
        # Filter by user
        user_id = self.request.GET.get('user')
        if user_id:
            queryset = queryset.filter(user_id=user_id)
        
        # Filter by action type
        action = self.request.GET.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by date range
        date_from = self.request.GET.get('date_from')
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        
        date_to = self.request.GET.get('date_to')
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'users': User.objects.all().order_by('username'),
            'action_choices': ['create', 'update', 'delete', 'movement_in', 'movement_out', 'assignment'],
        })
        return context

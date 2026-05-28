

import logging

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import models
from django.http import HttpResponse, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import redirect
from django.views import View
from django.views.generic import TemplateView, ListView
from django.utils import timezone
from django.contrib.auth import get_user_model
import csv
import io
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from attendance.exceptions import AttendanceServiceError
from attendance.models import Attendance
import attendance.leave_integration as attendance_services_module

check_in = attendance_services_module.check_in
check_out = attendance_services_module.check_out

logger = logging.getLogger(__name__)


class AttendanceCheckInView(LoginRequiredMixin, View):
    """Record check-in for the current user (POST only)."""

    http_method_names = ["post", "options"]

    def post(self, request, *args, **kwargs):
        try:
            check_in(request.user)
            messages.success(
                request,
                "You have checked in successfully.",
            )
        except AttendanceServiceError as exc:
            messages.error(request, str(exc))
        except DjangoValidationError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            logger.exception("Unexpected error during check-in")
            messages.error(
                request,
                f"Check-in error: {str(exc)}",
            )
        return redirect("attendance:attendance_dashboard")

    def dispatch(self, request, *args, **kwargs):
        if request.method.lower() == "get":
            return HttpResponseNotAllowed(["POST"])
        return super().dispatch(request, *args, **kwargs)


class AttendanceCheckOutView(LoginRequiredMixin, View):
    """Record check-out for the current user (POST only)."""

    http_method_names = ["post", "options"]

    def post(self, request, *args, **kwargs):
        try:
            check_out(request.user)
            messages.success(
                request,
                "You have checked out successfully.",
            )
        except AttendanceServiceError as exc:
            messages.error(request, str(exc))
        except DjangoValidationError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            logger.exception("Unexpected error during check-out")
            messages.error(
                request,
                f"Check-out error: {str(exc)}",
            )
        return redirect("attendance:attendance_dashboard")

    def dispatch(self, request, *args, **kwargs):
        if request.method.lower() == "get":
            return HttpResponseNotAllowed(["POST"])
        return super().dispatch(request, *args, **kwargs)


class TimesheetDownloadView(LoginRequiredMixin, View):
    """Generate and download timesheet for a given month/year."""
    
    def get(self, request, *args, **kwargs):
        try:
            # Get parameters
            month = int(request.GET.get('month', timezone.now().month))
            year = int(request.GET.get('year', timezone.now().year))
            format_type = request.GET.get('format', 'csv')
            employee_id = request.GET.get('employee', 'all')
            
            # Debug logging
            logger.info(f"Timesheet request: month={month}, year={year}, format={format_type}, employee={employee_id}")
            
            # Get attendance data
            User = get_user_model()
            from .models import Attendance
            
            # Filter employees
            if employee_id == 'all':
                if not (request.user.is_staff or getattr(request.user, 'role', None) in ['admin', 'manager']):
                    # Non-admin users can only download their own timesheet
                    employees = [request.user]
                else:
                    employees = User.objects.filter(is_active=True)
            else:
                employees = [User.objects.get(id=int(employee_id))]
            
            # Get attendance records for the month
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
            attendance_records = Attendance.objects.filter(
                date__gte=start_date,
                date__lte=end_date,
                user__in=employees
            ).select_related('user').order_by('date', 'user__username')
            
            logger.info(f"Found {len(attendance_records)} attendance records")
            
            # Generate timesheet based on format
            if format_type == 'csv':
                logger.info("Generating CSV timesheet")
                return self.generate_csv_timesheet(attendance_records, month, year)
            elif format_type == 'excel':
                logger.info("Generating Excel timesheet")
                return self.generate_excel_timesheet(attendance_records, month, year)
            else:  # pdf
                logger.info("Generating PDF timesheet")
                return self.generate_pdf_timesheet(attendance_records, month, year)
                
        except Exception as e:
            logger.exception(f"Error generating timesheet: {e}")
            return JsonResponse({'error': 'Failed to generate timesheet'}, status=500)
    
    def generate_csv_timesheet(self, records, month, year):
        """Generate CSV timesheet."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Timesheet Report', f'{month}/{year}'])
        writer.writerow([])
        writer.writerow(['Employee', 'Date', 'Check In', 'Check Out', 'Total Hours', 'Status'])
        
        # Data
        for record in records:
            status = 'Present' if record.check_in else 'Absent'
            if record.check_in and record.check_out:
                total_hours = record.total_hours or 0
            else:
                total_hours = 0
                
            writer.writerow([
                record.user.get_full_name() or record.user.username,
                record.date.strftime('%Y-%m-%d'),
                record.check_in.strftime('%H:%M') if record.check_in else '',
                record.check_out.strftime('%H:%M') if record.check_out else '',
                f'{total_hours:.1f}',
                status
            ])
        
        # Create response
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="timesheet_{month}_{year}.csv"'
        return response
    
    def generate_excel_timesheet(self, records, month, year):
        """Generate Excel timesheet with proper formatting."""
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Timesheet_{month}_{year}"  # Fixed: removed / character
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Timesheet Report - {month}/{year}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Employee', 'Date', 'Check In', 'Check Out', 'Total Hours', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        row_num = 4
        for record in records:
            # Determine status
            if record.check_in:
                if record.check_out:
                    status = 'Present'
                    total_hours = record.total_hours or 0
                else:
                    status = 'Checked In'
                    total_hours = 0
            else:
                status = 'Absent'
                total_hours = 0
            
            # Add data row
            ws.cell(row=row_num, column=1, value=record.user.get_full_name() or record.user.username)
            ws.cell(row=row_num, column=2, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=record.check_in.strftime('%H:%M') if record.check_in else 'N/A')
            ws.cell(row=row_num, column=4, value=record.check_out.strftime('%H:%M') if record.check_out else 'N/A')
            ws.cell(row=row_num, column=5, value=round(total_hours, 1))
            ws.cell(row=row_num, column=6, value=status)
            
            # Apply borders to data cells
            for col in range(1, 7):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center' if col in [2, 3, 4, 5] else 'left')
            
            # Color code status
            status_cell = ws.cell(row=row_num, column=6)
            if status == 'Present':
                status_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
            elif status == 'Absent':
                status_cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            elif status == 'Checked In':
                status_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            
            row_num += 1
        
        # Summary section
        summary_row = row_num + 2
        ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True)
        
        # Calculate summary statistics
        total_records = len(records)
        present_records = len([r for r in records if r.check_in])
        total_hours_sum = sum([r.total_hours or 0 for r in records])
        
        ws.cell(row=summary_row + 1, column=1, value='Total Records:')
        ws.cell(row=summary_row + 1, column=2, value=total_records)
        ws.cell(row=summary_row + 2, column=1, value='Present Days:')
        ws.cell(row=summary_row + 2, column=2, value=present_records)
        ws.cell(row=summary_row + 3, column=1, value='Total Hours:')
        ws.cell(row=summary_row + 3, column=2, value=round(total_hours_sum, 1))
        
        # Auto-adjust column widths (skip merged cells)
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells and get column letter from regular cells only
                    if cell.__class__.__name__ == 'MergedCell':
                        continue
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    if column_letter is None:
                        column_letter = cell.column_letter
                except:
                    pass
            if max_length > 0 and column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="timesheet_{month}_{year}.xlsx"'
        return response
    
    def generate_pdf_timesheet(self, records, month, year):
        """Generate PDF timesheet (simplified as text for now)."""
        output = io.StringIO()
        
        # Header
        output.write(f"Timesheet Report - {month}/{year}\n")
        output.write("=" * 50 + "\n\n")
        
        # Data
        for record in records:
            status = 'Present' if record.check_in else 'Absent'
            if record.check_in and record.check_out:
                total_hours = record.total_hours or 0
            else:
                total_hours = 0
                
            output.write(f"{record.user.get_full_name() or record.user.username}\n")
            output.write(f"  Date: {record.date.strftime('%Y-%m-%d')}\n")
            output.write(f"  Check In: {record.check_in.strftime('%H:%M') if record.check_in else 'N/A'}\n")
            output.write(f"  Check Out: {record.check_out.strftime('%H:%M') if record.check_out else 'N/A'}\n")
            output.write(f"  Total Hours: {total_hours:.1f}\n")
            output.write(f"  Status: {status}\n")
            output.write("-" * 30 + "\n")
        
        # Create response
        response = HttpResponse(output.getvalue(), content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="timesheet_{month}_{year}.txt"'
        return response


class CalendarExportView(LoginRequiredMixin, View):
    """Export calendar data for a given month/year."""
    
    def get(self, request, *args, **kwargs):
        try:
            # Get parameters
            month = int(request.GET.get('month', timezone.now().month))
            year = int(request.GET.get('year', timezone.now().year))
            
            # Get attendance data
            User = get_user_model()
            from .models import Attendance
            
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
            attendance_records = Attendance.objects.filter(
                date__gte=start_date,
                date__lte=end_date
            ).select_related('user').order_by('date', 'user__username')
            
            # Generate calendar data
            calendar_data = []
            current_date = start_date
            
            while current_date <= end_date:
                # Check if it's a weekend
                is_weekend = current_date.weekday() >= 5
                
                # Get attendance for this date
                day_attendance = attendance_records.filter(date=current_date)
                present_count = day_attendance.filter(check_in__isnull=False).count()
                total_employees = User.objects.filter(is_active=True).count()
                
                calendar_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': current_date.day,
                    'weekday': current_date.strftime('%A'),
                    'is_weekend': is_weekend,
                    'present_count': present_count,
                    'total_employees': total_employees,
                    'absent_count': total_employees - present_count
                })
                
                current_date += timedelta(days=1)
            
            # Generate text file instead of JSON
            output = io.StringIO()
            
            # Header
            output.write(f"Calendar Export - {month}/{year}\n")
            output.write("=" * 50 + "\n\n")
            
            # Calendar data
            for day in calendar_data:
                output.write(f"Date: {day['date']} ({day['weekday']})\n")
                output.write(f"  Present: {day['present_count']}/{day['total_employees']}\n")
                output.write(f"  Absent: {day['absent_count']}\n")
                output.write(f"  Weekend: {'Yes' if day['is_weekend'] else 'No'}\n")
                output.write("-" * 30 + "\n")
            
            # Summary statistics
            work_days = len([d for d in calendar_data if not d['is_weekend']])
            total_present = sum([d['present_count'] for d in calendar_data if not d['is_weekend']])
            avg_attendance = total_present / work_days if work_days > 0 else 0
            
            output.write(f"\nSummary Statistics\n")
            output.write("=" * 50 + "\n")
            output.write(f"Work Days: {work_days}\n")
            output.write(f"Total Present: {total_present}\n")
            output.write(f"Average Attendance: {avg_attendance:.1f}\n")
            output.write(f"Average Percentage: {(avg_attendance/20*100):.1f}%\n")  # Assuming 20 total employees
            
            # Create response
            response = HttpResponse(output.getvalue(), content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="calendar_{month}_{year}.txt"'
            return response
            
        except Exception as e:
            logger.exception(f"Error exporting calendar: {e}")
            return JsonResponse({'error': 'Failed to export calendar'}, status=500)


class AttendanceDashboardView(LoginRequiredMixin, TemplateView):
    """Attendance dashboard with today's performance and activity."""
    
    template_name = "attendance/dashboard.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        context["today"] = today
        
        # Import models here to avoid circular imports
        from .models import Attendance
        
        # Get the most recent attendance record for today
        context["today_attendance"] = (
            Attendance.objects.filter(user=self.request.user, date=today)
            .order_by('-check_in')
            .first()
        )
        
        # Check if user has an active check-in (checked in but not checked out)
        context["has_active_check_in"] = Attendance.objects.filter(
            user=self.request.user,
            date=today,
            check_out__isnull=True
        ).exists()
        
        # Add statistics for dashboard
        today_attendance_qs = Attendance.objects.filter(date=today)
        context["present_count"] = today_attendance_qs.filter(
            status__in=['present', 'late']
        ).count()
        context["absent_count"] = today_attendance_qs.filter(status='absent').count()
        context["late_count"] = today_attendance_qs.filter(status='late').count()
        context["total_employees"] = get_user_model().objects.filter(is_active=True).count()
        
        # Add work hour statistics for current user
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        
        user_week_attendance = Attendance.objects.filter(
            user=self.request.user,
            date__gte=week_start,
            date__lte=week_end
        )
        
        context["week_total_hours"] = user_week_attendance.aggregate(
            total_hours=models.Sum('total_hours')
        )['total_hours'] or 0
        context["week_overtime_hours"] = user_week_attendance.aggregate(
            total_overtime=models.Sum('overtime_hours')
        )['total_overtime'] or 0
        context["week_work_days"] = user_week_attendance.filter(
            check_in__isnull=False
        ).count()
        
        # Month statistics
        month_start = today.replace(day=1)
        user_month_attendance = Attendance.objects.filter(
            user=self.request.user,
            date__gte=month_start,
            date__lte=today
        )
        
        context["month_total_hours"] = user_month_attendance.aggregate(
            total_hours=models.Sum('total_hours')
        )['total_hours'] or 0
        context["month_overtime_hours"] = user_month_attendance.aggregate(
            total_overtime=models.Sum('overtime_hours')
        )['total_overtime'] or 0
        context["month_work_days"] = user_month_attendance.filter(
            check_in__isnull=False
        ).count()
        
        # Add recent attendance for current user
        context["recent_attendance"] = Attendance.objects.filter(
            user=self.request.user
        ).order_by('-date', '-check_in')[:10]
        
        # Pending approvals for managers/admins
        if self.request.user.is_staff or getattr(self.request.user, 'role', None) in ['admin', 'manager']:
            context["pending_approvals"] = Attendance.objects.filter(
                is_approved=False
            ).order_by('-date')[:10]
            context["has_pending_approvals"] = True
        else:
            context["pending_approvals"] = []
            context["has_pending_approvals"] = False
        
        return context


class CalendarDataView(LoginRequiredMixin, View):
    """Return calendar data as JSON for the modal display."""
    
    def get(self, request, *args, **kwargs):
        try:
            # Get parameters
            month = int(request.GET.get('month', timezone.now().month))
            year = int(request.GET.get('year', timezone.now().year))
            
            # Get attendance data
            User = get_user_model()
            from .models import Attendance
            
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
            
            attendance_records = Attendance.objects.filter(
                date__gte=start_date,
                date__lte=end_date
            ).select_related('user').order_by('date', 'user__username')
            
            # Generate calendar data
            calendar_data = []
            current_date = start_date
            
            while current_date <= end_date:
                # Check if it's a weekend
                is_weekend = current_date.weekday() >= 5
                
                # Get attendance for this date
                day_attendance = attendance_records.filter(date=current_date)
                present_count = day_attendance.filter(check_in__isnull=False).count()
                total_employees = User.objects.filter(is_active=True).count()
                
                calendar_data.append({
                    'date': current_date.strftime('%Y-%m-%d'),
                    'day': current_date.day,
                    'weekday': current_date.strftime('%A'),
                    'is_weekend': is_weekend,
                    'present_count': present_count,
                    'total_employees': total_employees,
                    'absent_count': total_employees - present_count
                })
                
                current_date += timedelta(days=1)
            
            # Return JSON response
            return JsonResponse({
                'month': month,
                'year': year,
                'calendar_data': calendar_data
            })
            
        except Exception as e:
            logger.exception(f"Error getting calendar data: {e}")
            return JsonResponse({'error': 'Failed to get calendar data'}, status=500)


class AttendanceTableView(LoginRequiredMixin, ListView):
    """Attendance table view with filters."""
    
    model = Attendance
    template_name = "attendance/attendance_table.html"
    context_object_name = "attendance_records"
    paginate_by = 30
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('user')
        
        # Filter by user (if not admin/manager, only show own attendance)
        user_id = self.request.GET.get('user')
        if not (self.request.user.is_staff or getattr(self.request.user, 'role', None) in ['admin', 'manager']):
            queryset = queryset.filter(user=self.request.user)
        elif user_id and user_id != 'all':
            queryset = queryset.filter(user_id=int(user_id))
        
        # Filter by date range
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            queryset = queryset.filter(date__lte=datetime.strptime(end_date, '%Y-%m-%d').date())
        
        # Filter by status
        status = self.request.GET.get('status')
        if status and status != 'all':
            queryset = queryset.filter(status=status)
        
        # Filter by approval status
        approved = self.request.GET.get('approved')
        if approved == 'pending':
            queryset = queryset.filter(is_approved=False)
        elif approved == 'approved':
            queryset = queryset.filter(is_approved=True)
        
        return queryset.order_by('-date', '-check_in')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = get_user_model().objects.filter(is_active=True).order_by('username')
        context['status_choices'] = Attendance.Status.choices
        return context


class AttendanceCalendarView(LoginRequiredMixin, TemplateView):
    """Calendar view for attendance."""
    
    template_name = "attendance/calendar.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        from .models import Attendance
        from calendar import monthcalendar, month_name
        
        # Handle month parameter - can be just month number or year-month format
        month_param = self.request.GET.get('month', '')
        if '-' in month_param:
            # Format: YYYY-MM
            year, month = map(int, month_param.split('-'))
        else:
            month = int(month_param) if month_param else today.month
            year = int(self.request.GET.get('year', today.year))
        
        context['month'] = month
        context['year'] = year
        context['month_name'] = month_name[month]
        context['today'] = today
        
        # Calculate previous/next month
        if month == 1:
            context['previous_month'] = 12
            context['previous_year'] = year - 1
        else:
            context['previous_month'] = month - 1
            context['previous_year'] = year
        
        if month == 12:
            context['next_month'] = 1
            context['next_year'] = year + 1
        else:
            context['next_month'] = month + 1
            context['next_year'] = year
        
        # Get employees visible on the calendar.
        selected_user_id = self.request.GET.get('user')
        can_view_all = self.request.user.is_staff or getattr(self.request.user, 'role', None) in ['admin', 'manager']
        all_users = get_user_model().objects.filter(is_active=True).order_by('username')
        if can_view_all:
            all_employees = all_users
            if selected_user_id:
                all_employees = all_employees.filter(id=selected_user_id)
        else:
            all_employees = all_users.filter(id=self.request.user.id)
            selected_user_id = str(self.request.user.id)

        context['employees'] = all_employees
        context['users'] = all_users if can_view_all else all_employees
        context['selected_user_id'] = int(selected_user_id) if selected_user_id else None
        
        # Get all attendance records for the month
        attendance_records = Attendance.objects.filter(
            date__year=year,
            date__month=month,
            user__in=all_employees,
        ).select_related('user')
        
        # Build a nested dictionary: attendance_dict[user_id][date] = attendance
        attendance_dict = {}
        for employee in all_employees:
            attendance_dict[employee.id] = {}
        
        for record in attendance_records:
            if record.user_id not in attendance_dict:
                attendance_dict[record.user_id] = {}
            attendance_dict[record.user_id][record.date] = record
        
        context['attendance_dict'] = attendance_dict
        
        # Build calendar data with all days
        calendar_days = []
        for week in monthcalendar(year, month):
            week_days = []
            for day in week:
                if day == 0:
                    week_days.append({'day': '', 'is_weekend': False})
                else:
                    date_obj = datetime(year, month, day).date()
                    is_weekend = date_obj.weekday() >= 5
                    is_future = date_obj > today
                    week_days.append({
                        'day': day,
                        'date': date_obj,
                        'is_weekend': is_weekend,
                        'is_future': is_future,
                        'is_absence_expected': not is_weekend and not is_future,
                    })
            calendar_days.append(week_days)
        
        context['calendar_days'] = calendar_days
        
        return context


class ManualAttendanceEntryView(LoginRequiredMixin, TemplateView):
    """Manual attendance entry form."""
    
    template_name = "attendance/manual_entry.html"
    
    def post(self, request, *args, **kwargs):
        from .models import Attendance
        from .forms import ManualAttendanceForm
        
        form = ManualAttendanceForm(request.POST)
        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.is_approved = False  # Requires approval
            attendance.save()
            
            # Recalculate hours
            attendance._recompute_total_hours()
            attendance.save()
            
            messages.success(request, "Attendance record created and pending approval.")
            return redirect("attendance:attendance_table")
        else:
            context = self.get_context_data()
            context['form'] = form
            return self.render_to_response(context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .forms import ManualAttendanceForm
        context['form'] = ManualAttendanceForm()
        context['users'] = get_user_model().objects.filter(is_active=True).order_by('username')
        return context


class ApproveAttendanceView(LoginRequiredMixin, View):
    """Approve or reject manual attendance entries."""
    
    def post(self, request, *args, **kwargs):
        from .models import Attendance
        
        attendance_id = request.POST.get('attendance_id')
        action = request.POST.get('action')  # 'approve' or 'reject'
        
        try:
            attendance = Attendance.objects.get(id=attendance_id)
            
            if action == 'approve':
                attendance.is_approved = True
                attendance.approved_by = request.user
                attendance.approved_at = timezone.now()
                attendance.save()
                messages.success(request, "Attendance record approved.")
            elif action == 'reject':
                attendance.delete()
                messages.success(request, "Attendance record rejected and deleted.")
            
        except Attendance.DoesNotExist:
            messages.error(request, "Attendance record not found.")
        
        return redirect("attendance:attendance_table")


class MonthlyReportView(LoginRequiredMixin, TemplateView):
    """Monthly report view with overtime details."""
    
    template_name = "attendance/monthly_report.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.localdate()
        
        # Handle month parameter - can be just month number or year-month format
        month_param = self.request.GET.get('month', '')
        if '-' in month_param:
            # Format: YYYY-MM
            year, month = map(int, month_param.split('-'))
        else:
            month = int(month_param) if month_param else today.month
            year = int(self.request.GET.get('year', today.year))
        
        context['month'] = month
        context['year'] = year
        
        # Get user filter
        user_id = self.request.GET.get('user')
        if user_id and (self.request.user.is_staff or getattr(self.request.user, 'role', None) in ['admin', 'manager']):
            from .models import Attendance
            user = get_user_model().objects.get(id=int(user_id))
            attendance_records = Attendance.objects.filter(
                user=user,
                date__year=year,
                date__month=month
            ).order_by('date')
            context['selected_user'] = user
        else:
            from .models import Attendance
            attendance_records = Attendance.objects.filter(
                user=self.request.user,
                date__year=year,
                date__month=month
            ).order_by('date')
            context['selected_user'] = self.request.user
        
        # Calculate statistics
        total_hours = attendance_records.aggregate(
            total=models.Sum('total_hours')
        )['total'] or 0
        total_overtime = attendance_records.aggregate(
            overtime=models.Sum('overtime_hours')
        )['overtime'] or 0
        present_days = attendance_records.filter(check_in__isnull=False).count()
        absent_days = attendance_records.filter(status='absent').count()
        late_days = attendance_records.filter(status='late').count()
        
        context['attendance_records'] = attendance_records
        context['total_hours'] = total_hours
        context['total_overtime'] = total_overtime
        context['present_days'] = present_days
        context['absent_days'] = absent_days
        context['late_days'] = late_days
        
        context['users'] = get_user_model().objects.filter(is_active=True).order_by('username')
        
        return context
